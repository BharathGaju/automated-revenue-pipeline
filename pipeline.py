# ============================================
# AUTOMATED REVENUE REPORTING PIPELINE
# Python + Google BigQuery + openpyxl
# ============================================

# ---- SETUP (run in Google Colab) -----------
# !pip install openpyxl google-cloud-bigquery pandas db-dtypes
# from google.colab import auth
# auth.authenticate_user()

import pandas as pd
import numpy as np
from google.cloud import bigquery
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ---- CONFIG --------------------------------
PROJECT_ID = "your-gcp-project-id"  # replace with your GCP project ID
client = bigquery.Client(project=PROJECT_ID)


# ---- STEP 1: PULL DATA FROM BIGQUERY -------
def fetch_data():
    query_weekly = """
        SELECT
            brand,
            DATE_TRUNC(order_date, WEEK)          AS week_start,
            COUNT(DISTINCT order_id)              AS total_orders,
            ROUND(SUM(grand_total), 2)            AS weekly_revenue,
            ROUND(AVG(grand_total), 2)            AS avg_order_value
        FROM `ecommerce_rfm.orders`
        WHERE status = 'complete'
        GROUP BY brand, week_start
        ORDER BY week_start, brand
    """
    query_monthly = """
        SELECT
            brand,
            DATE_TRUNC(order_date, MONTH)         AS month_start,
            COUNT(DISTINCT order_id)              AS total_orders,
            ROUND(SUM(grand_total), 2)            AS monthly_revenue,
            ROUND(AVG(grand_total), 2)            AS avg_order_value
        FROM `ecommerce_rfm.orders`
        WHERE status = 'complete'
        GROUP BY brand, month_start
        ORDER BY month_start, brand
    """
    query_summary = """
        SELECT
            brand,
            COUNT(DISTINCT order_id)              AS total_orders,
            ROUND(SUM(grand_total), 2)            AS total_revenue,
            ROUND(AVG(grand_total), 2)            AS avg_order_value,
            MIN(order_date)                       AS first_order,
            MAX(order_date)                       AS last_order
        FROM `ecommerce_rfm.orders`
        WHERE status = 'complete'
        GROUP BY brand
        ORDER BY total_revenue DESC
    """
    df_weekly  = client.query(query_weekly).to_dataframe()
    df_monthly = client.query(query_monthly).to_dataframe()
    df_summary = client.query(query_summary).to_dataframe()
    print(f"Data fetched — Weekly: {len(df_weekly)} | Monthly: {len(df_monthly)} | Summary: {len(df_summary)}")
    return df_summary, df_weekly, df_monthly


# ---- STEP 2: BUILD EXCEL REPORT ------------
def build_revenue_report(df_summary, df_weekly, df_monthly):

    wb = Workbook()

    # Styles
    header_font    = Font(bold=True, color="FFFFFF", size=11)
    header_fill    = PatternFill("solid", fgColor="1F4E79")
    alt_fill       = PatternFill("solid", fgColor="DEEAF1")
    center         = Alignment(horizontal="center", vertical="center")
    left           = Alignment(horizontal="left",   vertical="center")
    border_side    = Side(style="thin", color="BFBFBF")
    thin_border    = Border(left=border_side, right=border_side,
                            top=border_side,  bottom=border_side)

    def style_header_row(ws, row, num_cols):
        for col in range(1, num_cols + 1):
            cell            = ws.cell(row=row, column=col)
            cell.font       = header_font
            cell.fill       = header_fill
            cell.alignment  = center
            cell.border     = thin_border

    def style_data_row(ws, row, num_cols, alternate=False):
        for col in range(1, num_cols + 1):
            cell            = ws.cell(row=row, column=col)
            if alternate:
                cell.fill   = alt_fill
            cell.border     = thin_border
            cell.alignment  = left

    def set_col_widths(ws, widths):
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

    # ── SHEET 1: COVER ───────────────────────
    ws_cover = wb.active
    ws_cover.title = "Cover"
    ws_cover.sheet_view.showGridLines = False

    ws_cover.merge_cells("B2:F2")
    ws_cover["B2"].value     = "AUTOMATED REVENUE REPORT"
    ws_cover["B2"].font      = Font(bold=True, size=20, color="1F4E79")
    ws_cover["B2"].alignment = center

    ws_cover.merge_cells("B3:F3")
    ws_cover["B3"].value     = "Multi-Brand E-Commerce Analytics"
    ws_cover["B3"].font      = Font(size=13, color="2E75B6")
    ws_cover["B3"].alignment = center

    ws_cover.merge_cells("B5:F5")
    ws_cover["B5"].value     = f"Generated: {datetime.now().strftime('%B %d, %Y  %H:%M')}"
    ws_cover["B5"].font      = Font(size=11, italic=True, color="666666")
    ws_cover["B5"].alignment = center

    ws_cover.merge_cells("B6:F6")
    brands_str               = " | ".join(df_summary['brand'].tolist())
    ws_cover["B6"].value     = f"Brands: {brands_str}"
    ws_cover["B6"].font      = Font(size=11, color="444444")
    ws_cover["B6"].alignment = center

    kpis = [
        ("Total Revenue",   f"${df_summary['total_revenue'].sum():,.0f}"),
        ("Total Orders",    f"{df_summary['total_orders'].sum():,}"),
        ("Brands",          str(df_summary['brand'].nunique())),
        ("Avg Order Value", f"${df_summary['avg_order_value'].mean():,.2f}"),
    ]
    for i, (label, value) in enumerate(kpis):
        col = 2 + i
        ws_cover.merge_cells(start_row=8, start_column=col, end_row=8, end_column=col)
        ws_cover.merge_cells(start_row=9, start_column=col, end_row=9, end_column=col)
        lbl           = ws_cover.cell(row=8, column=col, value=label)
        val           = ws_cover.cell(row=9, column=col, value=value)
        lbl.font      = Font(bold=True, size=10, color="FFFFFF")
        lbl.fill      = PatternFill("solid", fgColor="1F4E79")
        lbl.alignment = center
        val.font      = Font(bold=True, size=14, color="1F4E79")
        val.fill      = PatternFill("solid", fgColor="DEEAF1")
        val.alignment = center
        ws_cover.row_dimensions[8].height = 25
        ws_cover.row_dimensions[9].height = 35
    set_col_widths(ws_cover, [3, 22, 22, 22, 22, 3])

    # ── SHEET 2: BRAND SUMMARY ───────────────
    ws_sum       = wb.create_sheet("Brand Summary")
    ws_sum.sheet_view.showGridLines = False
    ws_sum["A1"] = "BRAND REVENUE SUMMARY"
    ws_sum["A1"].font      = Font(bold=True, size=14, color="1F4E79")
    ws_sum.row_dimensions[1].height = 30

    headers = ["Brand", "Total Revenue ($)", "Total Orders",
               "Avg Order Value ($)", "First Order", "Last Order"]
    for col, h in enumerate(headers, 1):
        ws_sum.cell(row=2, column=col, value=h)
    style_header_row(ws_sum, 2, len(headers))

    for i, row in df_summary.iterrows():
        r = i + 3
        ws_sum.cell(row=r, column=1, value=row['brand'])
        ws_sum.cell(row=r, column=2, value=float(row['total_revenue'])).number_format = '#,##0.00'
        ws_sum.cell(row=r, column=3, value=int(row['total_orders']))
        ws_sum.cell(row=r, column=4, value=float(row['avg_order_value'])).number_format = '#,##0.00'
        ws_sum.cell(row=r, column=5, value=str(row['first_order']))
        ws_sum.cell(row=r, column=6, value=str(row['last_order']))
        style_data_row(ws_sum, r, len(headers), alternate=(i % 2 == 0))
    set_col_widths(ws_sum, [18, 22, 16, 22, 16, 16])

    # ── SHEET 3: MONTHLY REVENUE ─────────────
    ws_mon       = wb.create_sheet("Monthly Revenue")
    ws_mon.sheet_view.showGridLines = False
    ws_mon["A1"] = "MONTHLY REVENUE BY BRAND"
    ws_mon["A1"].font      = Font(bold=True, size=14, color="1F4E79")
    ws_mon.row_dimensions[1].height = 30

    mon_headers = ["Month", "Brand", "Monthly Revenue ($)", "Total Orders", "Avg Order Value ($)"]
    for col, h in enumerate(mon_headers, 1):
        ws_mon.cell(row=2, column=col, value=h)
    style_header_row(ws_mon, 2, len(mon_headers))

    df_monthly['month_start'] = pd.to_datetime(df_monthly['month_start'])
    for i, (_, row) in enumerate(df_monthly.sort_values(['month_start','brand']).iterrows()):
        r = i + 3
        ws_mon.cell(row=r, column=1, value=row['month_start'].strftime('%Y-%m'))
        ws_mon.cell(row=r, column=2, value=row['brand'])
        ws_mon.cell(row=r, column=3, value=float(row['monthly_revenue'])).number_format = '#,##0.00'
        ws_mon.cell(row=r, column=4, value=int(row['total_orders']))
        ws_mon.cell(row=r, column=5, value=float(row['avg_order_value'])).number_format = '#,##0.00'
        style_data_row(ws_mon, r, len(mon_headers), alternate=(i % 2 == 0))
    set_col_widths(ws_mon, [14, 18, 24, 16, 22])

    # ── SHEET 4: WEEKLY REVENUE ──────────────
    ws_wk        = wb.create_sheet("Weekly Revenue")
    ws_wk.sheet_view.showGridLines = False
    ws_wk["A1"] = "WEEKLY REVENUE BY BRAND"
    ws_wk["A1"].font      = Font(bold=True, size=14, color="1F4E79")
    ws_wk.row_dimensions[1].height = 30

    wk_headers = ["Week Start", "Brand", "Weekly Revenue ($)", "Total Orders", "Avg Order Value ($)"]
    for col, h in enumerate(wk_headers, 1):
        ws_wk.cell(row=2, column=col, value=h)
    style_header_row(ws_wk, 2, len(wk_headers))

    df_weekly['week_start'] = pd.to_datetime(df_weekly['week_start'])
    for i, (_, row) in enumerate(df_weekly.sort_values(['week_start','brand']).iterrows()):
        r = i + 3
        ws_wk.cell(row=r, column=1, value=row['week_start'].strftime('%Y-%m-%d'))
        ws_wk.cell(row=r, column=2, value=row['brand'])
        ws_wk.cell(row=r, column=3, value=float(row['weekly_revenue'])).number_format = '#,##0.00'
        ws_wk.cell(row=r, column=4, value=int(row['total_orders']))
        ws_wk.cell(row=r, column=5, value=float(row['avg_order_value'])).number_format = '#,##0.00'
        style_data_row(ws_wk, r, len(wk_headers), alternate=(i % 2 == 0))
    set_col_widths(ws_wk, [14, 18, 24, 16, 22])

    # ── SAVE ─────────────────────────────────
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    filename  = f"revenue_report_{timestamp}.xlsx"
    wb.save(filename)
    print(f"✅ Report saved: {filename}")
    print(f"   Sheets: {[s.title for s in wb.worksheets]}")
    return filename


# ---- STEP 3: RUN PIPELINE ------------------
if __name__ == "__main__":
    df_summary, df_weekly, df_monthly = fetch_data()
    filename = build_revenue_report(df_summary, df_weekly, df_monthly)
    print(f"\n🎉 Pipeline complete! Report: {filename}")
